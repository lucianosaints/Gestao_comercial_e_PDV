import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.styles import Font

def gerar_pdf_vendas(vendas):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "Relatório de Vendas - Patri-Tech")
    
    p.setFont("Helvetica", 12)
    y = 700
    p.drawString(100, y, "ID | Data | Valor Total | Forma Pagamento")
    y -= 20
    
    for venda in vendas:
        data_fmt = venda.data_venda.strftime('%d/%m/%Y %H:%M')
        p.drawString(100, y, f"{venda.id} | {data_fmt} | R$ {venda.valor_total} | {venda.forma_pagamento}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 750
            
    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer

def gerar_excel_inventario(bens):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário"
    
    headers = ["ID", "Nome", "Categoria", "Unidade", "Sala", "Valor", "Quantidade", "Situação"]
    ws.append(headers)
    
    for celula in ws[1]:
        celula.font = Font(bold=True)
        
    for bem in bens:
        ws.append([
            bem.id,
            bem.nome,
            bem.categoria.nome if bem.categoria else "-",
            bem.unidade.nome if bem.unidade else "-",
            bem.sala.nome if bem.sala else "-",
            float(bem.valor) if bem.valor else 0.0,
            bem.quantidade,
            bem.situacao
        ])
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_excel_despesas(despesas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Financeiro"
    
    headers = ["ID", "Descrição", "Categoria", "Valor (R$)", "Vencimento", "Status", "Fornecedor"]
    ws.append(headers)
    
    # Estilo do cabeçalho
    for celula in ws[1]:
        celula.font = Font(bold=True)
        
    for d in despesas:
        status = d.situacao
        venc = d.data_vencimento.strftime('%d/%m/%Y') if d.data_vencimento else "-"
        
        ws.append([
            d.id,
            d.descricao,
            d.get_tipo_display(), # Usa o nome legível ('Despesa Fixa' ao invés de 'FIXA')
            float(d.valor),
            venc,
            status,
            d.fornecedor.nome if d.fornecedor else "-"
        ])
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


from reportlab.graphics.barcode import code128
from reportlab.graphics import renderPDF
from reportlab.lib.units import mm

def gerar_etiquetas_pdf(bens):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter # 612 x 792 points (8.5 x 11 inch) aka 215.9 x 279.4 mm
    
    # Configuração da Etiqueta (Ex: 3 colunas x 8 linhas - A4 Seguro)
    # A4 = 210mm largura. 3 cols x 64mm + margens = 192mm (seguro)
    col_width = 64 * mm
    row_height = 33.9 * mm # Altura padrão Pimaco 6180
    
    # Margens para centralizar
    total_content_width = col_width * 3
    margin_x = (width - total_content_width) / 2
    margin_y = 12 * mm
    
    cols = 3
    rows = 8 # 8 linhas x 33.9mm = ~271mm (cabe em 297mm)
    
    x_offset = margin_x
    y_offset = height - margin_y - row_height
    
    col_counter = 0
    row_counter = 0
    
    for bem in bens:
        # Desenha a borda da etiqueta (opcional, bom para visualizar)
        c.setLineWidth(0.5)
        c.rect(x_offset, y_offset, col_width - 2*mm, row_height - 2*mm)
        
        # Conteúdo
        # 1. Nome do Produto (Truncado)
        c.setFont("Helvetica-Bold", 8)
        nome = bem.nome[:30] + "..." if len(bem.nome) > 30 else bem.nome
        c.drawString(x_offset + 5, y_offset + row_height - 15, nome) # Mais para baixo (era -10)
        
        # 2. Preço
        c.setFont("Helvetica-Bold", 10)
        preco = f"R$ {float(bem.valor):.2f}" if bem.valor else "R$ 0.00"
        c.drawString(x_offset + 5, y_offset + row_height - 27, preco) # Mais para baixo (era -22)
        
        # 3. Código de Barras
        codigo = bem.codigo_barras if bem.codigo_barras and bem.codigo_barras.strip() else f"ID-{bem.id}"
        barcode = code128.Code128(codigo, barHeight=8*mm, barWidth=0.8) # barHeight menor para caber texto
        
        # Centraliza o barcode na etiqueta
        barcode_x = x_offset + 5
        barcode_y = y_offset + 15 # Mais para cima (era +8) para dar espaço ao texto
        
        barcode.drawOn(c, barcode_x, barcode_y)

        # 4. Número do Código de Barras (Legível)
        c.setFont("Helvetica", 7)
        c.drawString(x_offset + 5, y_offset + 5, codigo)
        
        # Avança posição
        col_counter += 1
        x_offset += col_width
        
        if col_counter >= cols:
            col_counter = 0
            x_offset = margin_x
            row_counter += 1
            y_offset -= row_height
            
        if row_counter >= rows:
            c.showPage()
            row_counter = 0
            y_offset = height - margin_y - row_height
            col_counter = 0
            x_offset = margin_x
            
    c.save()
    buffer.seek(0)
    return buffer
